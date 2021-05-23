import openpyxl

def leer_celda(fila, columna):
    # Lee el archivo
    archivo = openpyxl.load_workbook('random.xlsx')
# Selecciona una hoja de trabajo
    hoja_activa = archivo.active
    # Selecciona una celda
    celda = hoja_activa.cell(fila, columna)
    # Asigna el valor de una celda
    valor = celda.value
    # Retorna el valor de esa celda
    return valor

def contar_filas():
    # Lee el archivo
    archivo = openpyxl.load_workbook('random.xlsx')
# Selecciona una hoja de trabajo
    hoja_activa = archivo.active
    # Selecciona una celda
    cantidad_filas = hoja_activa.max_row
    # Retorna el valor de esa celda
    return cantidad_filas
    

def contar_columnas():
    # Lee el archivo
    archivo = openpyxl.load_workbook('random.xlsx')
# Selecciona una hoja de trabajo
    hoja_activa = archivo.active
    # Selecciona una celda
    cantidad_columnas = hoja_activa.max_column
    # Retorna el valor de esa celda
    return cantidad_columnas

b2 = leer_celda(2, 2)
print(b2)
total_filas = contar_filas()
print(total_filas)
total_columnas = contar_columnas()
print (total_columnas)